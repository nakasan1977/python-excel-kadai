#coding:cp932

import pandas as pd
from glob import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# �t�@�C���p�X���擾
filepaths = glob("�f�[�^/*.xlsx")

# �f�[�^��ǂݍ��݁A�W�v
data = []
for filepath in filepaths:
    df = pd.read_excel(filepath, sheet_name="Sheet1")
    data.append(df)
combined_df = pd.concat(data, ignore_index=True,axis=0)
grouped_df = combined_df.groupby(["���i", "����N"]).agg({"���z�i��~�j": sum}).\
reset_index()

# ����W�v�\�쐬Excel�̍쐬
sales_totay = "����W�v�\.xlsx"
wb = Workbook()
ws = wb.active

# �W�v�f�[�^��Excel�ւ̓]�L
for row in dataframe_to_rows(combined_df, index=False, header=True):
    ws.append(row)

# �w�b�_�[�����̃Z���ɃO���[�̓h��Ԃ���K�p
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for cell in ws["1:1"]:
    cell.fill = header_fill

# Excel�t�@�C����ۑ�
wb.save(sales_totay)

# �t�@�C�������
wb.close()
